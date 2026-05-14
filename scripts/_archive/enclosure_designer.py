#!/usr/bin/env python3
"""
enclosure_designer.py – Conception de l'enclos stereo du SO-101

Genere un fichier Excel avec des tableaux explorant toutes les combinaisons
de 4 variables principales :
    D     : distance moteur du robot → cameras (cm)
    B     : baseline, ecart entre les 2 cameras stereo (cm)
    h     : hauteur des cameras au-dessus de la table (cm)
    theta : angle d'inclinaison des cameras sous l'horizontale (degres)

Les valeurs de focale et FOV sont lues directement depuis les fichiers
de calibration intrinseque (configs/calibration_cam_X.json).

Usage:
    python scripts/enclosure_designer.py
    python scripts/enclosure_designer.py --output mon_fichier.xlsx
"""

import json
import math
import os
import sys
import argparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl requis : pip install openpyxl")
    sys.exit(1)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  CONSTANTES DU ROBOT SO-101                                             ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

ARM_REACH = 43          # Portee du bras depuis le moteur de base (cm)
BASE_DEPTH = 7          # Profondeur de la base (dos de la base → moteur) (cm)
WORKSPACE_DIAMETER = 86  # Diametre du demi-cercle de travail = 2 * ARM_REACH (cm)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  VARIABLES PRINCIPALES – modifie ces valeurs pour explorer              ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

# Distance moteur → cameras (cm)
D_values = [55, 58, 60, 63, 65]

# Baseline stereo (cm)
B_values = [8, 10, 12, 15]

# Hauteur des cameras au-dessus de la table (cm)
h_values = [28, 30, 33, 35, 37, 40]

# Inclinaison sous l'horizontale (degres)
theta_values = [36, 38, 40, 42, 44, 46, 48, 50]


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  FORMULES                                                               ║
# ║                                                                         ║
# ║  Chaque fonction correspond a une formule numerotee.                    ║
# ║  Les docstrings expliquent la formule, ses variables et son usage.      ║
# ╚═══════════════════════════════════════════════════════════════════════════╝


def hfov_from_focal(fx, W=1920):
    """
    Formule 1 : HFOV = 2 * arctan( W / (2 * fx) )

    Champ de vision horizontal reel, calcule depuis la focale calibree.

        fx  = focale horizontale calibree (pixels)
        W   = largeur image (pixels), 1920 pour du 1080p
    """
    return 2 * math.degrees(math.atan(W / (2 * fx)))


def vfov_from_focal(fy, H=1080):
    """
    Formule 2 : VFOV = 2 * arctan( H / (2 * fy) )

    Champ de vision vertical reel, calcule depuis la focale calibree.

        fy  = focale verticale calibree (pixels)
        H   = hauteur image (pixels), 1080 pour du 1080p
    """
    return 2 * math.degrees(math.atan(H / (2 * fy)))


def coverage(Z, hfov_deg):
    """
    Formule 3 : L = 2 * Z * tan( HFOV / 2 )

    Largeur de scene visible a une distance Z de la camera.

        Z        = distance camera → ligne de scene (cm)
        hfov_deg = champ de vision horizontal (degres)
        L        = largeur couverte (cm)
    """
    return 2 * Z * math.tan(math.radians(hfov_deg / 2))


def x_near(h, theta_deg, vfov_deg):
    """
    Formule 4a : x_near = h / tan( theta + VFOV/2 )

    Point le plus PROCHE visible sur la table.
    Le bord inferieur du champ de vision touche la table a cette distance.

        h         = hauteur camera (cm)
        theta_deg = inclinaison sous l'horizontale (degres)
        vfov_deg  = champ de vision vertical (degres)
        x_near    = distance horizontale camera → point proche (cm)
    """
    angle = theta_deg + vfov_deg / 2
    if angle >= 90:
        return 0.0
    return h / math.tan(math.radians(angle))


def x_far(h, theta_deg, vfov_deg):
    """
    Formule 4b : x_far = h / tan( theta - VFOV/2 )

    Point le plus LOIN visible sur la table.
    Le bord superieur du champ de vision touche la table a cette distance.

        h         = hauteur camera (cm)
        theta_deg = inclinaison sous l'horizontale (degres)
        vfov_deg  = champ de vision vertical (degres)
        x_far     = distance horizontale camera → point loin (cm)
                     'inf' si le bord superieur est au-dessus de l'horizon
    """
    angle = theta_deg - vfov_deg / 2
    if angle <= 0:
        return float("inf")
    return h / math.tan(math.radians(angle))


def delta_z(Z, f, B):
    """
    Formule 5 : DeltaZ = Z^2 / ( f * B )

    Precision en profondeur du systeme stereo.
    C'est le plus petit ecart de profondeur mesurable (pour 1 pixel de disparite).

        Z  = distance camera → objet (cm)
        f  = focale en pixels
        B  = baseline en cm
        DeltaZ = resolution en profondeur (cm)

    Plus DeltaZ est petit, plus c'est precis.
    DeltaZ croit avec Z^2 : la precision se degrade vite avec la distance.
    """
    return Z ** 2 / (f * B)


def z_min(f, B, d_max=128):
    """
    Formule 6a : Z_min = ( f * B ) / d_max

    Distance minimale mesurable par l'algorithme stereo.

        f     = focale (pixels)
        B     = baseline (cm)
        d_max = disparite max de l'algorithme (pixels)
                128 ou 256 pour OpenCV StereoSGBM (param numDisparities)
    """
    return (f * B) / d_max


def z_max(f, B, d_min=1):
    """
    Formule 6b : Z_max = ( f * B ) / d_min

    Distance maximale mesurable. Au-dela, disparite < 1 pixel.

        d_min = disparite min detectable (1 px, ou 1/16 px en sub-pixel)
    """
    return (f * B) / d_min


def ratio_bz(B, Z):
    """
    Formule 7 : B/Z

    Ratio baseline / distance de travail.

    Plages de reference :
        >= 1/30       minimum (regle des stereographes)
        1/10 a 1/5    acceptable
        1/5  a 1/3    optimal (CalState, IEEE Kyto 2011)
        > 1/3         risque d'occlusion en environnement encombre
    """
    return B / Z


def relative_error(Z, f, B):
    """
    Formule 8 : DeltaZ / Z = Z / ( f * B )  =  1 / ( f * B/Z )

    Erreur relative en profondeur (sans unite, * 100 pour %).

    Plages de reference :
        < 1%    excellent (grasping fin)
        1-3%    bon (pick-and-place)
        > 5%    insuffisant pour manipulation precise
    """
    return Z / (f * B)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  LECTURE DES CALIBRATIONS                                                 ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

def load_calibrations(configs_dir="configs"):
    """Charge les 3 fichiers de calibration et retourne les parametres."""
    cams = {}
    for i in range(3):
        path = os.path.join(configs_dir, f"calibration_cam_{i}.json")
        if not os.path.exists(path):
            continue
        with open(path) as f:
            data = json.load(f)
        K = data["camera_matrix"]
        cams[i] = {
            "fx": K[0][0],
            "fy": K[1][1],
            "cx": K[0][2],
            "cy": K[1][2],
            "error": data["reprojection_error"],
            "captures": data["num_captures"],
        }
    return cams


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  STYLES EXCEL                                                             ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_BG = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FT = Font(bold=True, color="FFFFFF", size=10)
BORDER = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)


def write_header(ws, row, headers):
    for c, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = HEADER_FT
        cell.fill = HEADER_BG
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER


def write_cell(ws, row, col, value, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center")
    if fill:
        cell.fill = fill
    return cell


def autofit(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        lengths = [len(str(c.value or "")) for c in col]
        ws.column_dimensions[letter].width = min(max(lengths) + 3, 22)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  GENERATION EXCEL                                                         ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

def make_excel(cams, output_path):
    # Parametres stereo (moyenne cam_0 et cam_1)
    fx_avg = (cams[0]["fx"] + cams[1]["fx"]) / 2
    fy_avg = (cams[0]["fy"] + cams[1]["fy"]) / 2
    hfov = hfov_from_focal(fx_avg)
    vfov = vfov_from_focal(fy_avg)

    wb = Workbook()

    # ── ONGLET 1 : Cameras ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cameras"

    roles = {0: "stereo gauche", 1: "stereo droite", 2: "eye-in-hand"}
    headers = ["Camera", "Role", "fx (px)", "fy (px)", "HFOV (deg)", "VFOV (deg)",
               "Erreur (px)", "Captures"]
    write_header(ws, 1, headers)

    for r, (idx, cam) in enumerate(sorted(cams.items()), 2):
        vals = [
            f"cam_{idx}", roles[idx],
            round(cam["fx"], 1), round(cam["fy"], 1),
            round(hfov_from_focal(cam["fx"]), 1),
            round(vfov_from_focal(cam["fy"]), 1),
            round(cam["error"], 4), cam["captures"],
        ]
        for c, v in enumerate(vals, 1):
            fill = None
            if c == 7:  # erreur
                fill = GREEN if cam["error"] < 0.5 else YELLOW if cam["error"] < 1 else RED
            write_cell(ws, r, c, v, fill)

    r = 6
    ws.cell(row=r, column=1, value="Parametres stereo (moyenne cam_0 + cam_1) :").font = Font(bold=True)
    for i, line in enumerate([
        f"  fx moyen = {fx_avg:.1f} px",
        f"  fy moyen = {fy_avg:.1f} px",
        f"  HFOV = {hfov:.1f} deg",
        f"  VFOV = {vfov:.1f} deg",
    ]):
        ws.cell(row=r + 1 + i, column=1, value=line)
    autofit(ws)

    # ── ONGLET 2 : Precision stereo  (DeltaZ en fonction de B et D) ─────
    ws2 = wb.create_sheet("Precision stereo")

    ws2.cell(row=1, column=1, value="DeltaZ (mm) : precision en profondeur").font = Font(bold=True)
    ws2.cell(row=2, column=1, value=f"Formule 5 : DeltaZ = Z^2 / (f * B)    avec f = {fx_avg:.0f} px")

    # Pour chaque B, un tableau avec DeltaZ a differentes distances Z
    distances_Z = [15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65]
    row = 4
    headers = ["Z (cm)"] + [f"B = {B} cm" for B in B_values]
    write_header(ws2, row, headers)

    for Z in distances_Z:
        row += 1
        write_cell(ws2, row, 1, Z)
        for c, B in enumerate(B_values, 2):
            dz_mm = delta_z(Z, fx_avg, B) * 10  # cm → mm
            fill = GREEN if dz_mm < 2 else YELLOW if dz_mm < 5 else RED
            write_cell(ws2, row, c, round(dz_mm, 2), fill)

    row += 2
    ws2.cell(row=row, column=1, value="Vert < 2mm (grasping fin)  |  Jaune 2-5mm (pick-and-place)  |  Rouge > 5mm")

    # Sous-tableau : ratio B/Z
    row += 2
    ws2.cell(row=row, column=1, value="Ratio B/Z et erreur relative (%)").font = Font(bold=True)
    ws2.cell(row=row + 1, column=1, value="Formule 7 : B/Z    |    Formule 8 : erreur = Z / (f * B) * 100")
    row += 2
    write_header(ws2, row, ["Z (cm)"] + [f"B = {B} cm" for B in B_values])

    for Z in distances_Z:
        row += 1
        write_cell(ws2, row, 1, Z)
        for c, B in enumerate(B_values, 2):
            r_bz = ratio_bz(B, Z)
            err = relative_error(Z, fx_avg, B) * 100
            text = f"1/{Z / B:.0f}  ({err:.1f}%)"
            fill = GREEN if 1/5 <= r_bz <= 1/3 else YELLOW if 1/10 <= r_bz <= 1/2 else RED
            write_cell(ws2, row, c, text, fill)

    row += 2
    ws2.cell(row=row, column=1,
             value="Vert : B/Z entre 1/5 et 1/3 (optimal)  |  Jaune : 1/10-1/5 ou 1/3-1/2  |  Rouge : hors plage")

    # Sous-tableau : Z_min
    row += 2
    ws2.cell(row=row, column=1, value="Distance min/max mesurable").font = Font(bold=True)
    ws2.cell(row=row + 1, column=1, value="Formule 6a : Z_min = (f * B) / d_max")
    row += 2
    write_header(ws2, row, ["numDisparities"] + [f"B = {B} cm" for B in B_values])

    for d_max in [64, 128, 256]:
        row += 1
        write_cell(ws2, row, 1, f"d_max = {d_max}")
        for c, B in enumerate(B_values, 2):
            zm = z_min(fx_avg, B, d_max)
            write_cell(ws2, row, c, f"{zm:.1f} cm")

    autofit(ws2)

    # ── ONGLET 3 : Visibilite (x_near, x_far pour h et theta) ───────────
    ws3 = wb.create_sheet("Visibilite")

    ws3.cell(row=1, column=1, value="Zone visible sur la table pour chaque (h, theta)").font = Font(bold=True)
    ws3.cell(row=2, column=1, value=f"VFOV = {vfov:.1f} deg   |   Formules 4a et 4b")

    # x_near
    row = 4
    ws3.cell(row=row, column=1, value="x_near (cm) : point le plus proche visible").font = Font(bold=True)
    row += 1
    headers = ["theta \\ h"] + [f"h = {h} cm" for h in h_values]
    write_header(ws3, row, headers)

    for theta in theta_values:
        row += 1
        write_cell(ws3, row, 1, f"{theta} deg")
        for c, h in enumerate(h_values, 2):
            xn = x_near(h, theta, vfov)
            write_cell(ws3, row, c, round(xn, 1))

    # x_far
    row += 2
    ws3.cell(row=row, column=1, value="x_far (cm) : point le plus loin visible").font = Font(bold=True)
    row += 1
    write_header(ws3, row, headers)

    for theta in theta_values:
        row += 1
        write_cell(ws3, row, 1, f"{theta} deg")
        for c, h in enumerate(h_values, 2):
            xf = x_far(h, theta, vfov)
            val = "inf" if xf == float("inf") else round(xf, 1)
            write_cell(ws3, row, c, val)

    # Contraintes colorees pour chaque D
    for D in D_values:
        x_bras = D - ARM_REACH  # marge avec le bras
        row += 2
        ws3.cell(row=row, column=1,
                 value=f"Contrainte pour D = {D} cm  (bras a {x_bras} cm des cameras)").font = Font(bold=True)

        # x_near doit etre < x_bras
        row += 1
        ws3.cell(row=row, column=1, value=f"x_near < {x_bras} cm ?")
        row += 1
        write_header(ws3, row, ["theta \\ h"] + [f"h = {h} cm" for h in h_values])

        for theta in theta_values:
            row += 1
            write_cell(ws3, row, 1, f"{theta} deg")
            for c, h in enumerate(h_values, 2):
                xn = x_near(h, theta, vfov)
                ok = xn < x_bras
                write_cell(ws3, row, c, f"{xn:.1f}", GREEN if ok else RED)

        # x_far doit etre > D
        row += 1
        ws3.cell(row=row, column=1, value=f"x_far > {D} cm ?")
        row += 1
        write_header(ws3, row, ["theta \\ h"] + [f"h = {h} cm" for h in h_values])

        for theta in theta_values:
            row += 1
            write_cell(ws3, row, 1, f"{theta} deg")
            for c, h in enumerate(h_values, 2):
                xf = x_far(h, theta, vfov)
                ok = xf > D
                val = "inf" if xf == float("inf") else f"{xf:.1f}"
                write_cell(ws3, row, c, val, GREEN if ok else RED)

    autofit(ws3)

    # ── ONGLET 4 : Couverture laterale ──────────────────────────────────
    ws4 = wb.create_sheet("Couverture")

    ws4.cell(row=1, column=1, value="Largeur couverte (cm) a differentes distances").font = Font(bold=True)
    ws4.cell(row=2, column=1, value=f"Formule 3 : L = 2 * Z * tan(HFOV/2)    HFOV = {hfov:.1f} deg")
    ws4.cell(row=3, column=1, value=f"Workspace = {WORKSPACE_DIAMETER} cm (demi-cercle de rayon {ARM_REACH} cm)")

    row = 5
    headers = ["Z (cm)", "Largeur (cm)", f">= {WORKSPACE_DIAMETER} cm ?"]
    write_header(ws4, row, headers)

    for Z in [10, 15, 17, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65]:
        row += 1
        L = coverage(Z, hfov)
        ok = L >= WORKSPACE_DIAMETER
        write_cell(ws4, row, 1, Z)
        write_cell(ws4, row, 2, round(L, 1))
        write_cell(ws4, row, 3, "OUI" if ok else "NON", GREEN if ok else RED)

    autofit(ws4)

    # ── ONGLET 5 : Design final (toutes combinaisons) ───────────────────
    ws5 = wb.create_sheet("Design final")

    ws5.cell(row=1, column=1,
             value="Toutes les combinaisons (D, B, h, theta) avec verdicts").font = Font(bold=True)
    ws5.cell(row=2, column=1,
             value=f"f = {fx_avg:.0f} px  |  HFOV = {hfov:.1f} deg  |  VFOV = {vfov:.1f} deg  |  Bras = {ARM_REACH} cm")

    row = 4
    headers = [
        "D (cm)", "B (cm)", "h (cm)", "theta (deg)",
        "Enclos (cm)",
        "Marge bras (cm)",
        "x_near (cm)", "x_far (cm)",
        "Bras visible", "Robot visible",
        "DeltaZ centre (mm)", "DeltaZ robot (mm)",
        "B/Z centre", "Err. rel. (%)",
        "Couverture (cm)",
        "VERDICT",
    ]
    write_header(ws5, row, headers)

    for D in D_values:
        for B in B_values:
            for h in h_values:
                for theta in theta_values:
                    row += 1

                    enclos = D + BASE_DEPTH
                    marge = D - ARM_REACH
                    z_centre = D / 2  # centre du workspace

                    xn = x_near(h, theta, vfov)
                    xf = x_far(h, theta, vfov)

                    bras_ok = xn < marge
                    robot_ok = (xf > D) if xf != float("inf") else True

                    dz_centre = delta_z(z_centre, fx_avg, B) * 10  # mm
                    dz_robot = delta_z(D, fx_avg, B) * 10  # mm

                    bz = ratio_bz(B, z_centre)
                    err = relative_error(z_centre, fx_avg, B) * 100

                    cov = coverage(D, hfov)
                    cov_ok = cov >= WORKSPACE_DIAMETER

                    tout_ok = bras_ok and robot_ok and cov_ok and dz_centre < 5

                    values = [
                        D, B, h, theta,
                        enclos,
                        marge,
                        round(xn, 1),
                        "inf" if xf == float("inf") else round(xf, 1),
                        "OUI" if bras_ok else "NON",
                        "OUI" if robot_ok else "NON",
                        round(dz_centre, 2),
                        round(dz_robot, 2),
                        f"1/{z_centre / B:.0f}",
                        round(err, 2),
                        round(cov, 1),
                        "OK" if tout_ok else "",
                    ]

                    for c, v in enumerate(values, 1):
                        fill = None
                        if c == 9:   fill = GREEN if bras_ok else RED
                        if c == 10:  fill = GREEN if robot_ok else RED
                        if c == 11:  fill = GREEN if dz_centre < 2 else YELLOW if dz_centre < 5 else RED
                        if c == 12:  fill = GREEN if dz_robot < 5 else YELLOW if dz_robot < 10 else RED
                        if c == 15:  fill = GREEN if cov_ok else RED
                        if c == 16:  fill = GREEN if tout_ok else None
                        write_cell(ws5, row, c, v, fill)

    autofit(ws5)

    # Sauvegarder
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  MAIN                                                                     ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

def main():
    parser = argparse.ArgumentParser(description="Conception de l'enclos stereo SO-101")
    parser.add_argument("--output", default="outputs/enclosure_design.xlsx", help="Fichier Excel de sortie")
    parser.add_argument("--configs-dir", default="configs", help="Dossier des calibrations")
    args = parser.parse_args()

    cams = load_calibrations(args.configs_dir)

    print("=== Cameras chargees ===")
    roles = {0: "stereo gauche", 1: "stereo droite", 2: "eye-in-hand"}
    for idx, cam in sorted(cams.items()):
        print(f"  cam_{idx} ({roles[idx]}): fx={cam['fx']:.1f}  fy={cam['fy']:.1f}  erreur={cam['error']:.4f}px")

    fx_avg = (cams[0]["fx"] + cams[1]["fx"]) / 2
    fy_avg = (cams[0]["fy"] + cams[1]["fy"]) / 2
    print(f"\n=== Parametres stereo (cam_0 + cam_1) ===")
    print(f"  fx moyen = {fx_avg:.1f} px")
    print(f"  HFOV     = {hfov_from_focal(fx_avg):.1f} deg")
    print(f"  VFOV     = {vfov_from_focal(fy_avg):.1f} deg")

    print(f"\n=== Variables explorees ===")
    print(f"  D     = {D_values} cm")
    print(f"  B     = {B_values} cm")
    print(f"  h     = {h_values} cm")
    print(f"  theta = {theta_values} deg")
    total = len(D_values) * len(B_values) * len(h_values) * len(theta_values)
    print(f"  → {total} combinaisons")

    path = make_excel(cams, args.output)
    print(f"\n=== Fichier genere : {path} ===")
    print("Ouvre-le et filtre la colonne VERDICT sur 'OK' dans l'onglet 'Design final'.")


if __name__ == "__main__":
    main()
