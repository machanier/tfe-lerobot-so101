#!/usr/bin/env python3
"""Campagne experimentale de mesures pick-and-place.

Produit un jeu de mesures reproductibles pour la section "Resultats" du
chapitre 6 du memoire. Pour chaque position d'objet, le script lance N essais
et collecte, par essai :
  - success (bool)     : le robot a-t-il saisi l'objet ?
  - n_attempts (int)   : nombre de tentatives (1 = succes direct, 2 = avec retry).
  - duration_s (float) : duree totale du pick-and-place.
  - attempts_log (list): pince reelle, marge et succes par tentative.

Sorties :
  - outputs/experiments/campaign_YYYYMMDD_HHMMSS.json : toutes les donnees.
  - outputs/experiments/campaign_YYYYMMDD_HHMMSS.csv  : vue tabulaire.
  - Statistiques agregees affichees a l'ecran (taux de succes, temps moyen).

Protocole :
  1. Placer la boite de depose et renseigner sa position (configs/scene.json).
  2. Choisir les positions d'objet (par exemple : centre, gauche, droite).
  3. Lancer le script avec --n-per-position et --positions.
  4. A chaque essai, le script demande de replacer l'objet a la position
     indiquee, puis attend ENTREE.
  5. Le robot tente la saisie ; les resultats sont journalises et sauvegardes.
  6. En fin de campagne, les statistiques et les fichiers de sortie sont
     recapitules.

Usage typique :
  python scripts/experiment_campaign.py \\
      --target orange_cube --detector hf \\
      --n-per-position 10 \\
      --positions centre gauche droite

A executer avec le robot sous tension et l'ensemble des peripheriques (cameras
et hub USB) en place. Verifier au prealable que
`python scripts/check_calibration.py` passe.
"""

import argparse
import csv
import json
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO))
sys.path.insert(0, str(REPO / "scripts"))

from config import FOLLOWER_PORT  # noqa: E402
from src.pipeline import PickAndPlacePipeline, PipelineConfig  # noqa: E402


def make_pipeline(args) -> PickAndPlacePipeline:
    """Cree un nouveau pipeline pour un essai (etat reset)."""
    cfg = PipelineConfig(
        target_label=args.target,
        detector_kind=args.detector,
        motor_port=args.port,
        max_velocity_rad_s=args.max_velocity,
        grip_close_pct=args.grip_close,
        dry_run=False,                      # campagne reelle
        closed_loop=True,                   # actif pour une comparaison equitable
        display=args.display,
        max_grasp_retries=args.max_retries,
    )
    # Seuil de detection de la saisie : par defaut, herite du defaut calibre de
    # PipelineConfig (8.0 pour le cube 30 mm). Il n'est surcharge que si
    # --grasp-threshold est explicitement fourni (analyse de sensibilite).
    if args.grasp_threshold is not None:
        cfg.grasp_success_threshold_pct = args.grasp_threshold
    if args.grasp_lateral_offset is not None:
        cfg.grasp_lateral_offset_mm = args.grasp_lateral_offset
    return PickAndPlacePipeline(cfg)


def _ask_manual_observations(record: dict, args) -> None:
    """Demande les observations manuelles juste apres l'essai.

    Les reponses sont ecrites directement dans le CSV et le classeur Excel.
    ENTREE valide la valeur par defaut. L'orientation est reportee d'un essai a
    l'autre : elle n'est ressaisie que lorsqu'elle change. Le mode d'echec n'est
    demande que si l'essai a echoue. Ctrl+C saute les questions ; pour arreter la
    campagne, utiliser Ctrl+C a la demande de placement suivante.
    """
    success = bool(record.get("success"))
    try:
        last_or = getattr(args, "_last_orientation", "") or ""
        prompt = (f"   orientation ? [ENTREE = {last_or}] : " if last_or
                  else "   orientation ? [ex. : debout/couche/a plat, ENTREE = rien] : ")
        o = input(prompt).strip()
        orientation = o if o else last_or
        args._last_orientation = orientation
        record["orientation"] = orientation

        auto = "oui" if success else "non"
        d = input(f"   depose dans la boite ? [ENTREE = {auto}, ou o/n] : ").strip().lower()
        record["depose_boite"] = ("oui" if d in ("o", "oui", "y")
                                  else "non" if d in ("n", "non") else auto)

        if not success or record["depose_boite"] == "non":
            m = input("   mode d'echec ? [E1 E2 E3 E4 E5 E6, ENTREE = rien] : ").strip().upper()
            record["mode_echec"] = m if m in {"E1", "E2", "E3", "E4", "E5", "E6"} else ""
        else:
            record["mode_echec"] = ""

        record["notes"] = input("   note ? [ENTREE = rien] : ").strip()
    except (EOFError, KeyboardInterrupt):
        print("   (questions sautees)")


def run_trial(args, position_name: str, trial_idx: int) -> dict:
    """Execute un essai et renvoie un dictionnaire de metriques."""
    print()
    print(">" * 70)
    print(f">>> Essai {trial_idx} a la position '{position_name}'")
    print(">" * 70)
    if not args.no_prompt:
        try:
            input(f"    Placer l'objet a la position '{position_name}' "
                  f"puis ENTREE (Ctrl+C pour arreter la campagne) : ")
        except (EOFError, KeyboardInterrupt):
            raise

    t0 = time.time()
    record = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "position": position_name,
        "trial": trial_idx,
        "success": None,
        "n_attempts": None,
        "attempts_log": [],
        "duration_s": None,
        "grasp_pose_error_mm": None,   # residu IK de la prise retenue (erreur de pose atteinte)
        "cam2_correction_mm": None,    # amplitude de la correction cam_2 (precision de perception)
        "orientation": None,           # saisie manuelle
        "depose_boite": None,          # saisie manuelle
        "mode_echec": None,            # saisie manuelle
        "notes": None,                 # saisie manuelle
        "error": None,
    }
    try:
        pipeline = make_pipeline(args)
        pipeline.run()
        record["success"] = bool(getattr(pipeline, "_grasp_final_success", False))
        record["n_attempts"] = int(getattr(pipeline, "_grasp_total_attempts", 0))
        record["attempts_log"] = list(getattr(pipeline, "_grasp_attempts_log", []))
        # Metriques de precision internes : None si jamais renseignees (par
        # exemple echec de perception avant planification, ou absence de
        # correction cam_2).
        record["grasp_pose_error_mm"] = getattr(pipeline, "_grasp_pose_error_mm", None)
        record["cam2_correction_mm"] = getattr(pipeline, "_cam2_correction_mm", None)
    except KeyboardInterrupt:
        record["error"] = "interrupted"
        record["success"] = False
        raise   # remonte a l'appelant pour un arret propre
    except Exception as e:
        record["error"] = f"{type(e).__name__}: {e}"
        record["success"] = False
        print(f"    !! Exception : {record['error']}")
        traceback.print_exc()
    finally:
        record["duration_s"] = round(time.time() - t0, 2)

    tag = "OK" if record["success"] else "ECHEC"
    n = record["n_attempts"] or 0
    print(f"    >>> {position_name} essai {trial_idx} : {tag} "
          f"en {n} tentative(s), {record['duration_s']:.1f}s")
    if not args.no_prompt:
        _ask_manual_observations(record, args)
    return record


def print_aggregate_stats(results: list[dict]):
    """Affiche les statistiques agregees des essais."""
    print()
    print("=" * 70)
    print(" Statistiques agregees (chapitre 6)")
    print("=" * 70)
    if not results:
        print("  Aucun resultat (essais interrompus avant toute execution).")
        return

    # Par position
    by_pos: dict[str, list[dict]] = {}
    for r in results:
        by_pos.setdefault(r.get("position", "?"), []).append(r)

    for pos, trials in by_pos.items():
        n = len(trials)
        ok = sum(1 for t in trials if t.get("success"))
        rate = 100.0 * ok / n if n else 0.0
        # Tentatives moyennes parmi les succes (1.0 si tous directs, environ 1.5
        # si la moitie a demande un retry).
        succ_attempts = [t.get("n_attempts") or 1 for t in trials if t.get("success")]
        avg_attempts = (sum(succ_attempts) / len(succ_attempts)) if succ_attempts else float("nan")
        durations = [t.get("duration_s") or 0 for t in trials if t.get("duration_s")]
        avg_dur = (sum(durations) / len(durations)) if durations else 0.0
        # Essais ayant utilise au moins un retry (n_attempts > 1).
        n_with_retry = sum(1 for t in trials if (t.get("n_attempts") or 0) > 1)
        print(f"  {pos:<15} : {ok:>2}/{n:>2} OK ({rate:>5.1f}%)"
              f"  tentatives moyennes sur succes = {avg_attempts:.2f}"
              f"  temps moyen = {avg_dur:.1f}s"
              f"  retries utilises = {n_with_retry}/{n}")

    # Global
    n_total = len(results)
    n_ok = sum(1 for r in results if r.get("success"))
    rate_g = 100.0 * n_ok / n_total if n_total else 0.0
    durations = [r.get("duration_s") or 0 for r in results if r.get("duration_s")]
    avg_dur_g = (sum(durations) / len(durations)) if durations else 0.0
    n_retry_used = sum(1 for r in results if (r.get("n_attempts") or 0) > 1)
    n_retry_saved = sum(1 for r in results
                        if r.get("success") and (r.get("n_attempts") or 0) > 1)
    print()
    print(f"  GLOBAL   : {n_ok:>2}/{n_total:>2} OK ({rate_g:.1f}%)  "
          f"temps moyen = {avg_dur_g:.1f}s")
    print(f"  RETRY    : {n_retry_used} essais avec retry,"
          f" dont {n_retry_saved} rattrapes grace au retry")
    print("=" * 70)


def save_results(results: list[dict], args, out_dir: Path) -> tuple[Path, Path]:
    """Sauvegarde les resultats en JSON (complet) et en CSV (vue tabulaire)."""
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = out_dir / f"campaign_{stamp}.json"
    csv_path = out_dir / f"campaign_{stamp}.csv"

    with open(json_path, "w") as f:
        json.dump({
            "args": vars(args),
            "n_results": len(results),
            "results": results,
        }, f, indent=2, default=str)

    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        target = getattr(args, "target", "")
        detector = getattr(args, "detector", "")
        writer.writerow(["objet", "detecteur",
                         "timestamp", "position", "trial", "success",
                         "n_attempts", "duration_s",
                         "grasp_pose_error_mm", "cam2_correction_mm",
                         "last_gripper_pct", "last_margin_pct",
                         "error",
                         # colonnes a renseigner manuellement dans le tableur :
                         "orientation", "depose_boite", "mode_echec", "notes"])
        for r in results:
            last = r["attempts_log"][-1] if r.get("attempts_log") else {}
            gpe = r.get("grasp_pose_error_mm")
            c2c = r.get("cam2_correction_mm")
            writer.writerow([
                target,
                detector,
                r.get("timestamp"),
                r.get("position"),
                r.get("trial"),
                int(bool(r.get("success"))),
                r.get("n_attempts"),
                f"{r.get('duration_s') or 0:.2f}",
                f"{gpe:.2f}" if gpe is not None else "",
                f"{c2c:.2f}" if c2c is not None else "",
                f"{last.get('gripper_pct', ''):.1f}" if last.get("gripper_pct") is not None else "",
                f"{last.get('marge_pct', ''):+.1f}" if last.get("marge_pct") is not None else "",
                r.get("error") or "",
                r.get("orientation") or "", r.get("depose_boite") or "",
                r.get("mode_echec") or "", r.get("notes") or "",
            ])
    return json_path, csv_path


def append_to_xlsx(results: list[dict], args, xlsx_path=None) -> None:
    """Ajoute les essais dans l'onglet 'Donnees' du classeur de suivi.

    Les lignes sont ecrites dans docs/campagne_suivi.xlsx a la fin de chaque
    commande ; la synthese se recalcule a l'ouverture. Les colonnes manuelles
    (orientation, depose_boite, mode_echec, notes) restent vides et sont a
    renseigner a la main.

    La fonction est tolerante : si le classeur est absent, ouvert dans un tableur
    ou si openpyxl est manquant, rien n'est ecrit (le CSV reste la reference) et
    la raison est affichee. Elle ne leve pas d'exception.
    """
    xlsx_path = Path(xlsx_path) if xlsx_path else (REPO / "docs" / "campagne_suivi.xlsx")
    if not xlsx_path.exists():
        print(f"   [xlsx] {xlsx_path.name} introuvable ; rien ajoute (le CSV suffit).")
        return

    def _num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        if "Donnees" not in wb.sheetnames:
            print("   [xlsx] onglet 'Donnees' absent ; rien ajoute (le CSV suffit).")
            return
        ws = wb["Donnees"]
        target = getattr(args, "target", "")
        detector = getattr(args, "detector", "")
        for r in results:
            last = r["attempts_log"][-1] if r.get("attempts_log") else {}
            ws.append([
                target, detector,
                r.get("timestamp"), r.get("position"), r.get("trial"),
                int(bool(r.get("success"))), r.get("n_attempts"),
                _num(r.get("duration_s")),
                _num(r.get("grasp_pose_error_mm")), _num(r.get("cam2_correction_mm")),
                _num(last.get("gripper_pct")), _num(last.get("marge_pct")),
                r.get("error") or "",
                r.get("orientation") or "", r.get("depose_boite") or "",
                r.get("mode_echec") or "", r.get("notes") or "",
            ])
        wb.save(xlsx_path)
        print(f"   [xlsx] {len(results)} essai(s) ajoute(s) dans {xlsx_path.name} "
              f"(onglet Donnees) ; renseigner les colonnes manuelles.")
    except PermissionError:
        print(f"   [xlsx] {xlsx_path.name} est ouvert dans un tableur ; le fermer "
              f"et relancer, ou coller le CSV a la main (le CSV est deja sauvegarde).")
    except Exception as e:
        print(f"   [xlsx] non ecrit ({type(e).__name__}: {e}) ; utiliser le CSV.")


def main():
    p = argparse.ArgumentParser(
        description="Campagne experimentale pick-and-place (TFE chapitre 6).",
    )
    p.add_argument("--target", default="orange_cube",
                   help="Label de l'objet a saisir (defaut : orange_cube).")
    p.add_argument("--detector", default="hf", choices=["hsv", "hf"],
                   help="Detecteur a utiliser (defaut : hf = OWL-ViTv2).")
    p.add_argument("--n-per-position", type=int, default=10,
                   help="Nombre d'essais par position (defaut : 10).")
    p.add_argument("--positions", nargs="+",
                   default=["centre", "gauche", "droite"],
                   help="Noms logiques des positions (defaut : centre gauche droite).")
    p.add_argument("--port", default=FOLLOWER_PORT,
                   help="Port USB du follower (defaut : valeur de config.FOLLOWER_PORT).")
    p.add_argument("--max-velocity", type=float, default=0.5,
                   help="Vitesse articulaire maximale en rad/s (defaut : 0.5).")
    p.add_argument("--grip-close", type=float, default=5.0,
                   help="Consigne de fermeture de la pince, de 0 a 100 "
                        "(defaut : 5, soit presque fermee).")
    p.add_argument("--max-retries", type=int, default=1,
                   help="Nombre maximal de retry par essai "
                        "(defaut : 1, soit 2 tentatives au total).")
    p.add_argument("--grasp-threshold", type=float, default=None,
                   help="Seuil de detection de la saisie, en points de pourcentage "
                        "au-dessus de grip-close. Par defaut, herite du defaut calibre "
                        "de PipelineConfig (8.0 pour le cube 30 mm). A ne forcer que "
                        "pour une analyse de sensibilite.")
    p.add_argument("--grasp-lateral-offset", type=float, default=None,
                   help="Decalage lateral de la saisie en mm (pince asymetrique). "
                        "Par defaut, 8 (cube 30 mm). A regler par objet : un rectangle "
                        "ou un cylindre plus large peut demander une autre valeur.")
    p.add_argument("--output-dir", default="outputs/experiments",
                   help="Repertoire de sortie (defaut : outputs/experiments).")
    p.add_argument("--no-prompt", action="store_true",
                   help="N'attend pas ENTREE entre les essais (mode automatique, "
                        "adapte a une alimentation continue des objets).")
    p.add_argument("--display", action="store_true",
                   help="Affiche le flux des cameras pendant chaque essai "
                        "(ralentit l'execution).")
    args = p.parse_args()

    print("=" * 70)
    print(" Campagne experimentale pick-and-place")
    print("=" * 70)
    print(f"  cible            : {args.target}")
    print(f"  detecteur        : {args.detector}")
    print(f"  positions        : {args.positions}")
    print(f"  essais/position  : {args.n_per_position}")
    print(f"  total essais     : {args.n_per_position * len(args.positions)}")
    print(f"  max retries      : {args.max_retries}")
    seuil_txt = (f"{args.grasp_threshold}%" if args.grasp_threshold is not None
                 else "8% (defaut calibre de PipelineConfig)")
    print(f"  seuil saisie OK  : pince > consigne + {seuil_txt}")
    print("=" * 70)
    print()
    if not args.no_prompt:
        try:
            input("Pret a commencer la campagne ? ENTREE = demarrer, Ctrl+C = annuler : ")
        except (EOFError, KeyboardInterrupt):
            print("\nAnnule.")
            return

    out_dir = REPO / args.output_dir
    results: list[dict] = []
    try:
        for pos_name in args.positions:
            print()
            print("#" * 70)
            print(f"#  Position : {pos_name}  ({args.n_per_position} essais)")
            print("#" * 70)
            for trial in range(1, args.n_per_position + 1):
                record = run_trial(args, pos_name, trial)
                results.append(record)
    except KeyboardInterrupt:
        print("\n!! Campagne interrompue par l'utilisateur (Ctrl+C).")
    except Exception as e:
        print(f"\n!! Erreur fatale : {type(e).__name__}: {e}")
        traceback.print_exc()
    finally:
        # Sauvegarde systematique des resultats disponibles, meme en cas
        # d'interruption en cours de route.
        if results:
            json_path, csv_path = save_results(results, args, out_dir)
            print()
            print(f">> {len(results)} essais enregistres :")
            print(f"   {json_path}")
            print(f"   {csv_path}")
            append_to_xlsx(results, args)
            print_aggregate_stats(results)
        else:
            print("Aucun resultat a sauvegarder.")


if __name__ == "__main__":
    main()
